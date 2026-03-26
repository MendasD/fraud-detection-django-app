"""
apps/dashboard/views.py
Vues du tableau de bord analytique Fortal Bank.
"""

import json
from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from django.http import JsonResponse
from django.db.models import Count, Sum, Avg, Q
from django.views import View

from apps.transactions.models import Transaction, Alert


class DashboardIndexView(LoginRequiredMixin, TemplateView):
    """Vue principale du dashboard — KPIs + graphiques + carte + alertes live."""
    template_name = 'dashboard/index.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        last_24h = now - timedelta(hours=24)
        last_7d  = now - timedelta(days=7)
        last_30d = now - timedelta(days=30)

        # KPIs principaux
        total_txn      = Transaction.objects.count()
        txn_24h        = Transaction.objects.filter(timestamp__gte=last_24h).count()
        fraud_statuses = ['SUSPECTE', 'BLOQUEE']
        total_fraud    = Transaction.objects.filter(status__in=fraud_statuses).count()
        fraud_24h      = Transaction.objects.filter(timestamp__gte=last_24h, status__in=fraud_statuses).count()
        total_amount   = Transaction.objects.aggregate(s=Sum('amount'))['s'] or 0
        fraud_amount   = Transaction.objects.filter(status__in=fraud_statuses).aggregate(s=Sum('amount'))['s'] or 0
        pending_alerts = Alert.objects.filter(status='NOUVELLE').count()
        avg_score      = Transaction.objects.filter(fraud_score__isnull=False).aggregate(a=Avg('fraud_score'))['a'] or 0

        # Taux de fraude
        fraud_rate = (total_fraud / total_txn * 100) if total_txn > 0 else 0

        # Données pour graphiques (JSON) 
        # Transactions par heure (dernières 24h)
        txn_by_hour = self._get_txn_by_hour(last_24h)

        # Répartition par type
        txn_by_type = list(
            Transaction.objects.values('transaction_type')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )

        # Transactions par ville (top 10)
        txn_by_city = list(
            Transaction.objects.values('city')
            .annotate(count=Count('id'), fraud_count=Count('id', filter=Q(status__in=fraud_statuses)))
            .order_by('-count')[:10]
        )

        # Évolution fraudes sur 30 jours
        fraud_trend = self._get_fraud_trend(last_30d)

        # Alertes récentes (20 dernières)
        recent_alerts = Alert.objects.select_related('transaction').filter(
            status='NOUVELLE'
        ).order_by('-created_at')[:20]

        ctx.update({
            # KPIs
            'total_txn':       total_txn,
            'txn_24h':         txn_24h,
            'total_fraud':     total_fraud,
            'fraud_24h':       fraud_24h,
            'fraud_rate':      round(fraud_rate, 2),
            'total_amount':    int(total_amount),
            'fraud_amount':    int(fraud_amount),
            'pending_alerts':  pending_alerts,
            'avg_score':       round(avg_score * 100, 1),
            'txn_by_city':     txn_by_city,
            # Données graphiques (sérialisées en JSON pour le JS)
            'txn_by_hour_json':  json.dumps(txn_by_hour),
            'txn_by_type_json':  json.dumps(txn_by_type),
            'txn_by_city_json':  json.dumps(txn_by_city),
            'fraud_trend_json':  json.dumps(fraud_trend),
            # Valeurs numériques JS (via json.dumps pour éviter la localisation fr : "10,55" ou "3 000")
            'js_kpis': json.dumps({
                'total_txn':     total_txn,
                'txn_24h':       txn_24h,
                'total_fraud':   total_fraud,
                'fraud_24h':     fraud_24h,
                'fraud_rate':    round(fraud_rate, 2),
                'total_amount':  int(total_amount),
                'fraud_amount':  int(fraud_amount),
                'pending_alerts': pending_alerts,
                'avg_score':     round(avg_score * 100, 1),
            }),
            # Alertes
            'recent_alerts':   recent_alerts,
        })
        return ctx

    def _get_txn_by_hour(self, since):
        """Retourne 24 points horaires fixes depuis 'since', avec 0 pour les heures vides."""
        from django.db.models.functions import TruncHour
        qs = (
            Transaction.objects
            .filter(timestamp__gte=since)
            .annotate(hour=TruncHour('timestamp'))
            .values('hour')
            .annotate(count=Count('id'), fraud=Count('id', filter=Q(status__in=['SUSPECTE', 'BLOQUEE'])))
            .order_by('hour')
        )
        by_hour = {item['hour']: item for item in qs if item['hour']}
        result = []
        now = timezone.now()
        start = since.replace(minute=0, second=0, microsecond=0)
        for h in range(25):
            slot = start + timedelta(hours=h+1)
            if slot.date() < now.date():
                label = slot.strftime('Hier %H:%M')
            else:
                label = slot.strftime('%H:%M')
            item = by_hour.get(slot)
            result.append({
                'hour':  label,
                'count': item['count'] if item else 0,
                'fraud': item['fraud'] if item else 0,
            })
        return result[:24]

    def _get_fraud_trend(self, since):
        """Retourne le nombre de fraudes par jour sur les 30 derniers jours."""
        from django.db.models.functions import TruncDate
        qs = (
            Transaction.objects
            .filter(timestamp__gte=since, status__in=['SUSPECTE', 'BLOQUEE'])
            .annotate(date=TruncDate('timestamp'))
            .values('date')
            .annotate(count=Count('id'))
            .order_by('date')
        )
        return [
            {'date': item['date'].strftime('%d/%m'), 'count': item['count']}
            for item in qs
        ]


class TransactionListView(LoginRequiredMixin, TemplateView):
    """Liste des transactions avec filtres."""
    template_name = 'dashboard/transactions.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = Transaction.objects.all().order_by('-timestamp')

        # Filtres
        status = self.request.GET.get('status')
        city   = self.request.GET.get('city')
        txn_type = self.request.GET.get('type')

        if status:
            qs = qs.filter(status=status)
        if city:
            qs = qs.filter(city=city)
        if txn_type:
            qs = qs.filter(transaction_type=txn_type)

        ctx['transactions'] = qs[:200]
        ctx['cities']       = Transaction.objects.values_list('city', flat=True).distinct().order_by('city')
        ctx['status_choices'] = Transaction.Status.choices
        ctx['type_choices']   = Transaction.TransactionType.choices
        return ctx


class AlertListView(LoginRequiredMixin, TemplateView):
    """Liste de toutes les alertes."""
    template_name = 'dashboard/alerts.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['alerts'] = Alert.objects.select_related('transaction', 'resolved_by').order_by('-created_at')[:100]
        return ctx


class MapView(LoginRequiredMixin, TemplateView):
    """Vue carte des transactions géolocalisées au Sénégal."""
    template_name = 'dashboard/map.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Transactions avec coordonnées GPS
        txns = Transaction.objects.filter(
            location_lat__isnull=False,
            location_lon__isnull=False,
        ).values(
            'transaction_id', 'amount', 'transaction_type',
            'location_lat', 'location_lon', 'city',
            'status', 'fraud_score', 'timestamp'
        ).order_by('-timestamp')[:500]

        ctx['transactions_geo_json'] = json.dumps([
            {
                'id':    str(t['transaction_id']),
                'lat':   t['location_lat'],
                'lon':   t['location_lon'],
                'amount': int(t['amount']),
                'type':  t['transaction_type'],
                'city':  t['city'],
                'status': t['status'],
                'score': round(t['fraud_score'] or 0, 2),
                'time':  t['timestamp'].strftime('%d/%m %H:%M') if t['timestamp'] else '',
            }
            for t in txns
        ])
        return ctx


# ── API JSON pour les mises à jour dynamiques ────────────────────────────────

class StatsAPIView(LoginRequiredMixin, View):
    """Endpoint JSON pour les statistiques live (polling fallback)."""

    def get(self, request):
        now = timezone.now()
        last_24h = now - timedelta(hours=24)

        total_txn   = Transaction.objects.count()
        total_fraud = Transaction.objects.filter(status__in=['SUSPECTE', 'BLOQUEE']).count()
        amounts     = Transaction.objects.aggregate(
            total_amount=Sum('amount'),
            fraud_amount=Sum('amount', filter=Q(status__in=['SUSPECTE', 'BLOQUEE'])),
            avg_score=Avg('fraud_score'),
        )

        stats = {
            'total_txn':      total_txn,
            'total_fraud':    total_fraud,
            'fraud_rate':     round(total_fraud / total_txn * 100, 1) if total_txn else 0,
            'txn_24h':        Transaction.objects.filter(timestamp__gte=last_24h).count(),
            'fraud_24h':      Transaction.objects.filter(timestamp__gte=last_24h, status__in=['SUSPECTE', 'BLOQUEE']).count(),
            'pending_alerts': Alert.objects.filter(status='NOUVELLE').count(),
            'total_amount':   int(amounts['total_amount'] or 0),
            'fraud_amount':   int(amounts['fraud_amount'] or 0),
            'avg_score':      round((amounts['avg_score'] or 0) * 100, 1),
            'last_alert':     None,
        }

        last = Alert.objects.filter(status='NOUVELLE').order_by('-created_at').first()
        if last:
            stats['last_alert'] = {
                'id':     last.id,
                'level':  last.level,
                'score':  round(last.fraud_score, 2),
                'amount': int(last.transaction.amount),
                'city':   last.transaction.city,
                'time':   last.created_at.strftime('%H:%M:%S'),
            }

        return JsonResponse(stats)


class ExportTransactionsView(LoginRequiredMixin, View):
    """Export des transactions en CSV ou Excel."""

    COLUMNS = [
        ('transaction_id', 'ID Transaction'),
        ('timestamp',      'Date/Heure'),
        ('sender_name',    'Expéditeur'),
        ('receiver_name',  'Destinataire'),
        ('amount',         'Montant (FCFA)'),
        ('transaction_type', 'Type'),
        ('city',           'Ville'),
        ('status',         'Statut'),
        ('fraud_score',    'Score Fraude'),
        ('device_type',    'Appareil'),
    ]

    def get(self, request):
        fmt = request.GET.get('format', 'csv').lower()
        qs = Transaction.objects.values(*[c[0] for c in self.COLUMNS]).order_by('-timestamp')
        headers = [c[1] for c in self.COLUMNS]
        fields  = [c[0] for c in self.COLUMNS]

        if fmt == 'excel':
            return self._export_excel(qs, headers, fields)
        return self._export_csv(qs, headers, fields)

    def _export_csv(self, qs, headers, fields):
        import csv
        from django.http import StreamingHttpResponse

        def rows():
            yield headers
            for row in qs:
                yield [str(row.get(f, '') or '') for f in fields]

        class Echo:
            def write(self, value): return value

        writer = csv.writer(Echo())
        response = StreamingHttpResponse(
            (writer.writerow(r) for r in rows()),
            content_type='text/csv; charset=utf-8-sig',
        )
        response['Content-Disposition'] = 'attachment; filename="transactions_fortal.csv"'
        return response

    def _export_excel(self, qs, headers, fields):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        # En-têtes
        header_fill = PatternFill(fill_type='solid', fgColor='00C853')
        header_font = Font(bold=True, color='000000')
        for col, label in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Données
        for row_idx, row in enumerate(qs, 2):
            for col_idx, field in enumerate(fields, 1):
                import uuid
                value = row.get(field, '')
                if hasattr(value, 'strftime'):
                    value = value.strftime('%d/%m/%Y %H:%M')
                elif isinstance(value, uuid.UUID):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Largeurs de colonnes
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from django.http import HttpResponse as HR
        response = HR(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="transactions_fortal.xlsx"'
        return response
